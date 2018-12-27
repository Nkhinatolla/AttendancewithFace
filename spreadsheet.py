from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time
import os
import sqlite3

#database connection
conn = sqlite3.connect('Face-DataBase')
c = conn.cursor()

#get current date
currentDate = time.strftime("%d_%m_%y")

#create a workbook and add a worksheet
if(os.path.exists('./reports.xlsx')):
    wb = load_workbook(filename = "reports.xlsx")
    sheet = wb.get_sheet_by_name('Cse15')
    # sheet[ord() + '1']
    for col_index in range(1, 100):
        col = get_column_letter(col_index)
        if sheet['%s%s' % (col,1)].value is None:
            col2 = get_column_letter(col_index - 1)
            if sheet['%s%s' % (col2,1)].value != currentDate:
                sheet['%s%s' % (col,1)].value = currentDate
            break
    attend = [0 for i in range(201)]
    r = 3
    first = False
    for row_index in range(2, 200):
        x = sheet['A%s' % row_index].value
        if x is not None:
            attend[x] = True
        else:
            if (first == False):
                first = True
                r = row_index
    print('Last None: ' + str(r))
    c.execute("SELECT * FROM Students ORDER BY ID ASC")
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            if (attend[a[0]] != True):
                print('Cordinate : ' + ('A%s' % r) + str(a[0]))
                sheet['A%s' % r].value = a[0]
                sheet['B%s' % r].value = a[1]
                r += 1
    #saving the file
    wb.save(filename = "reports.xlsx")
else:
    wb = Workbook()
    dest_filename = 'reports.xlsx'
    c.execute("SELECT * FROM Students ORDER BY ID ASC")

    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "Cse15"
    ws1.append(('ID', 'Name', currentDate))

    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[0], a[1]))

    #saving the file
    wb.save(filename = dest_filename)
