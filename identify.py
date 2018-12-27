import cognitive_face as CF
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time


#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "reports.xlsx")
sheet = wb.get_sheet_by_name('Cse15')

def getDateColumn():
    y = sheet.rows
    y = list(y)
    for i in range(1, len(y[0]) + 1):
        col = get_column_letter(i)
        if sheet['%s%s'% (col,'1')].value == currentDate:
            return col


Key = 'd45ebcdf3bb8479980a4224f0944a24d'
CF.Key.set(Key)
BASE_URL = 'https://westeurope.api.cognitive.microsoft.com/face/v1.0/'  # Replace with your regional Base URL
CF.BaseUrl.set(BASE_URL)

connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(100)]

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
    if filename.endswith(".jpg"):
        imgurl = urllib.pathname2url(os.path.join(directory, filename))
        print(imgurl)
        print("-----------")
        res = CF.face.detect(imgurl)
        if len(res) != 1:
        	print ("No face detected.")
        	continue

        faceIds = []
        for face in res:
        	faceIds.append(face['faceId'])
        res = CF.face.identify(faceIds, "test1")
        print (res)
        for face in res:
            if not face['candidates']:
            	print ("Unknown")
            else:
                personId = face['candidates'][0]['personId']
                print(personId)
                c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
                row = c.fetchone()
                if (row is not None):
                    attend[row[0]] = 1
                    print (row[1] + " recognized")
                else:
                    print("There is some problem in DB!!!")
x = sheet.columns
x = list(x)
for row in range(2, len(x[0]) + 1):
    rn = sheet['A%s' % row].value
    if rn is not None:
        col = getDateColumn()
        sheet['%s%s' % (col, str(row))].value = attend[rn]

wb.save(filename = "reports.xlsx")
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,"test1")
# for face in res:
#     personName = CF.person.get("test1", face['candidates']['personId'])
#     print personName
#print res
